"""ExcelExporter — escribe xlsx con filas discrepantes resaltadas rojo."""
from __future__ import annotations
import logging
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

log = logging.getLogger(__name__)

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(color="9C0006", bold=True)
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


class ExcelExporter:
    def __init__(self, output_path: Path):
        self.output_path = output_path

    def exportar(self, df: pd.DataFrame, sheet_name: str = "Cruce") -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        flag_cols = ["__SIN_RETENCION", "__DIFF_MONTO"]
        flags_df = df[flag_cols].copy() if all(c in df.columns for c in flag_cols) else None

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT

        # pintar filas con flag
        if flags_df is not None:
            for i, (sin_ret, diff) in enumerate(flags_df.itertuples(index=False), start=2):
                if bool(sin_ret) or bool(diff):
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=i, column=c).fill = RED_FILL
                        ws.cell(row=i, column=c).font = RED_FONT

        # auto-width simple
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        ws.freeze_panes = "A2"
        wb.save(self.output_path)
        log.info(f"xlsx exportado: {self.output_path}")
        return self.output_path
