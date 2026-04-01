from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.orm import Session
from pydantic import BaseModel, EmailStr
from typing import List, Optional
import io, csv

from app.api.deps import get_db, get_current_user, get_current_admin, get_current_admin_or_supervisor
from app.models.empleado import Empleado, EmpleadoEstado, Grupo
from app.models.user import User, RolEnum

router = APIRouter()

# ── Schemas ────────────────────────────────────────────────────────────────────

class EmpleadoCreate(BaseModel):
    cedula: str
    nombre: str
    apellido: str
    email: Optional[str] = None
    departamento: Optional[str] = None
    cargo: Optional[str] = None
    grupo_id: Optional[int] = None

class EmpleadoUpdate(BaseModel):
    nombre: Optional[str] = None
    apellido: Optional[str] = None
    email: Optional[str] = None
    departamento: Optional[str] = None
    cargo: Optional[str] = None
    grupo_id: Optional[int] = None
    estado: Optional[str] = None

class EmpleadoResponse(BaseModel):
    id: int
    cedula: str
    nombre: str
    apellido: str
    email: Optional[str]
    departamento: Optional[str]
    cargo: Optional[str]
    estado: str
    grupo_id: Optional[int]
    grupo_nombre: Optional[str] = None

    class Config:
        from_attributes = True

class ImportResult(BaseModel):
    insertados: int
    omitidos: int
    errores: List[str]

# ── Helper ─────────────────────────────────────────────────────────────────────

def _enriquecer(emp: Empleado, db: Session) -> EmpleadoResponse:
    grupo_nombre = None
    if emp.grupo_id:
        g = db.query(Grupo).filter(Grupo.id == emp.grupo_id).first()
        grupo_nombre = g.nombre if g else None
    return EmpleadoResponse(
        id=emp.id, cedula=emp.cedula, nombre=emp.nombre, apellido=emp.apellido,
        email=emp.email, departamento=emp.departamento, cargo=emp.cargo,
        estado=emp.estado.value, grupo_id=emp.grupo_id, grupo_nombre=grupo_nombre
    )

def _filtrar_por_rol(query, current_user: User):
    """Operador solo ve empleados de sus grupos. Admin/Supervisor ve todos."""
    if current_user.rol == RolEnum.operador:
        from app.models.empleado import OperadorGrupo
        ids_grupos = [a.grupo_id for a in
                      db.query(OperadorGrupo).filter(OperadorGrupo.user_id == current_user.id).all()]
        if not ids_grupos:
            return query.filter(False)
        return query.filter(Empleado.grupo_id.in_(ids_grupos))
    return query

# ── GET /empleados/ ────────────────────────────────────────────────────────────

@router.get("/", response_model=List[EmpleadoResponse])
def get_empleados(
    grupo_id: Optional[int] = Query(None),
    estado: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_or_supervisor)
):
    """Lista empleados con filtros opcionales. Admin y Supervisor."""
    q = db.query(Empleado)
    if grupo_id:
        q = q.filter(Empleado.grupo_id == grupo_id)
    if estado:
        try:
            q = q.filter(Empleado.estado == EmpleadoEstado(estado))
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Estado inválido: {estado}")
    return [_enriquecer(e, db) for e in q.order_by(Empleado.apellido).all()]

# ── POST /empleados/ (crear uno) ───────────────────────────────────────────────

@router.post("/", response_model=EmpleadoResponse, status_code=201)
def create_empleado(
    payload: EmpleadoCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_or_supervisor)
):
    """Crea un empleado individual. Admin y Supervisor."""
    if db.query(Empleado).filter(Empleado.cedula == payload.cedula).first():
        raise HTTPException(status_code=400, detail=f"Ya existe un empleado con cédula '{payload.cedula}'.")
    emp = Empleado(**payload.model_dump(), estado=EmpleadoEstado.prueba)
    db.add(emp); db.commit(); db.refresh(emp)
    return _enriquecer(emp, db)

# ── PATCH /empleados/{id} ──────────────────────────────────────────────────────

@router.patch("/{empleado_id}", response_model=EmpleadoResponse)
def update_empleado(
    empleado_id: int,
    payload: EmpleadoUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_or_supervisor)
):
    """Actualiza datos de un empleado. Admin y Supervisor."""
    emp = db.query(Empleado).filter(Empleado.id == empleado_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Empleado no encontrado.")
    for field, val in payload.model_dump(exclude_none=True).items():
        if field == "estado":
            try:
                setattr(emp, field, EmpleadoEstado(val))
            except ValueError:
                raise HTTPException(status_code=400, detail=f"Estado inválido: {val}")
        else:
            setattr(emp, field, val)
    db.commit(); db.refresh(emp)
    return _enriquecer(emp, db)

# ── DELETE /empleados/{id} ─────────────────────────────────────────────────────

@router.delete("/{empleado_id}")
def delete_empleado(
    empleado_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_or_supervisor)
):
    """Elimina un empleado. Admin y Supervisor."""
    emp = db.query(Empleado).filter(Empleado.id == empleado_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Empleado no encontrado.")
    db.delete(emp); db.commit()
    return {"msg": f"Empleado '{emp.nombre} {emp.apellido}' eliminado."}

# ── POST /empleados/importar ───────────────────────────────────────────────────

COLUMNAS_REQUERIDAS = {"cedula", "nombre", "apellido"}
COLUMNAS_OPCIONALES = {"email", "departamento", "cargo", "grupo_id"}

@router.post("/importar", response_model=ImportResult)
async def importar_empleados(
    file: UploadFile = File(...),
    grupo_id: Optional[int] = Query(None, description="Si se pasa, todos los empleados importados se asignan a este grupo"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_or_supervisor)
):
    """
    Importa empleados desde un archivo Excel (.xlsx) o CSV (.csv).
    Columnas requeridas: cedula, nombre, apellido
    Columnas opcionales: email, departamento, cargo, grupo_id
    Los empleados duplicados (misma cédula) se omiten sin error.
    """
    filename = file.filename.lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".csv")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx o .csv")

    contenido = await file.read()
    filas: List[dict] = []

    if filename.endswith(".xlsx"):
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl no instalado. Ejecuta: pip install openpyxl")
        wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        ws = wb.active
        headers = []
        for row in ws.iter_rows(values_only=True):
            row_vals = [str(c).strip().lower().replace(" ", "_") if c is not None else "" for c in row]
            if "nombre" in row_vals and "apellido" in row_vals:
                headers = row_vals
                break
        
        # Correccion para plantillas donde 'cedula' esta en columna A y los datos reales en B
        if len(headers) >= 2 and headers[0] == "cedula" and headers[1] == "":
            headers[0] = "ignore"
            headers[1] = "cedula"

        if not headers:
            wb.close()
            raise HTTPException(status_code=400, detail="No se encontraron las columnas 'nombre' y 'apellido' en el Excel.")

        # Leer los datos despues de la cabecera
        is_data = False
        for row in ws.iter_rows(values_only=True):
            row_vals = [str(c).strip().lower() if c is not None else "" for c in row]
            if not is_data:
                if "nombre" in row_vals and "apellido" in row_vals:
                    is_data = True
                continue
            
            if any(row):  # Ignorar filas totalmente vacías
                filas.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row) if i < len(headers)})
        wb.close()
    else:
        text = contenido.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        filas = [{k.strip().lower(): (v.strip() if v else "") for k, v in row.items()} for row in reader]

    # Validar que las columnas requeridas existan
    if not filas:
        raise HTTPException(status_code=400, detail="El archivo está vacío.")
    cabeceras = set(filas[0].keys())
    faltantes = COLUMNAS_REQUERIDAS - cabeceras
    if faltantes:
        raise HTTPException(status_code=400, detail=f"Columnas requeridas faltantes: {', '.join(faltantes)}")

    insertados, omitidos, errores = 0, 0, []

    # Obtener grupos válidos para evitar errores de Foreign Key
    from app.models.empleado import Grupo
    grupos_validos = {g[0] for g in db.query(Grupo.id).all()}

    for i, fila in enumerate(filas, start=2):
        # Omitir filas que están completamente vacías o solo tienen espacios en el Excel
        if not any(str(v).strip() for v in fila.values()):
            continue

        cedula = fila.get("cedula", "").strip()
        nombre = fila.get("nombre", "").strip()
        apellido = fila.get("apellido", "").strip()

        if not cedula or not nombre or not apellido:
            errores.append(f"Fila {i}: cedula, nombre y apellido son obligatorios.")
            continue

        # Omitir duplicados silenciosamente
        if db.query(Empleado).filter(Empleado.cedula == cedula).first():
            omitidos += 1
            continue

        # Resolver grupo_id: prioridad → columna del archivo → parámetro URL
        gid_raw = fila.get("grupo_id", "")
        gid = None
        if gid_raw:
            try:
                parsed_gid = int(gid_raw)
                if parsed_gid in grupos_validos:
                    gid = parsed_gid
            except ValueError:
                pass
        
        if gid is None and grupo_id:
            gid = grupo_id

        emp = Empleado(
            cedula=cedula,
            nombre=nombre,
            apellido=apellido,
            email=fila.get("email") or None,
            departamento=fila.get("departamento") or None,
            cargo=fila.get("cargo") or None,
            grupo_id=gid,
            estado=EmpleadoEstado.prueba,  # Pendiente hasta que se active en Prompt 8
        )
        db.add(emp)
        insertados += 1

    db.commit()
    return ImportResult(insertados=insertados, omitidos=omitidos, errores=errores)
